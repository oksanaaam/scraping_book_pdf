import re
import openpyxl
import fitz

from openpyxl.styles import Font


pdf_file = "C:/Users/User/projects/scraping_book_pdf/Abstract Book.pdf"
xl_file = "C:/Users/User/projects/scraping_book_pdf/task.xlsx"

start_page = 43
last_page = 64


class Article:
    def __init__(self):
        self.session_name = ""
        self.session_title = ""
        self.authors = []
        self.affiliations = []
        self.presentation_abstract = ""
        self.location = []

    def __str__(self):
        return (
            f"\nSession name: {self.session_name}"
            f"\nSession title: {self.session_title}"
            f"\nAuthors: {', '.join(self.authors)}"
            f"\nAffiliations: {', '.join(self.affiliations)}"
            f"\nLocation: {', '.join(self.location)}"
            f"\nPresentation abstract: {self.presentation_abstract}"
        )


def extract_articles_from_pdf(file_path: str):
    articles = []
    current_article = Article()
    previous_article = None
    previous_author = None

    file = fitz.open(file_path)

    for page_number in range(start_page - 1, last_page):
        page = file[page_number]
        text_blocks = page.get_text("dict")["blocks"]

        for block in text_blocks:
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"]

                    authors = (
                        span["font"] == "TimesNewRomanPS-ItalicMT" and span["size"] == 9
                    )
                    affiliations_data = (
                        span["font"] == "TimesNewRomanPS-ItalicMT" and span["size"] == 8
                    )
                    abstract_data = span["size"] == 9.134002685546875

                    session_name_match = re.match(r"P\d+", text)
                    if session_name_match:
                        current_article.session_name = session_name_match.group(
                            0
                        ).strip()

                    if current_article.session_name and text.isupper():
                        session_title = re.sub(
                            r"^" + re.escape(current_article.session_name),
                            "",
                            text.strip(),
                            flags=re.IGNORECASE,
                        )
                        current_article.session_title += session_title.strip()

                    elif authors:
                        author = text.replace("- ", "").strip()
                        if author:
                            if previous_author and author[0].islower():
                                current_article.authors[-1] += " " + author
                            else:
                                current_article.authors.append(author)
                            previous_author = author
                    elif affiliations_data:
                        if "," in text:
                            words = text.split(",")
                            affiliations = []
                            location = ""
                            for word in words:
                                word = word.strip()
                                if word:
                                    if len(word.split()) == 1:
                                        location += word + ", "
                                    else:
                                        affiliations.append(word)
                            location = location.rstrip(", ")
                            if location:
                                current_article.location.append(location)
                            current_article.affiliations.extend(affiliations)
                        else:
                            if text.strip():
                                current_article.affiliations.append(text)

                    elif abstract_data:
                        current_article.presentation_abstract += text

            if block["type"] == 0:
                if (
                    current_article.session_name
                    or current_article.session_title
                    or current_article.authors
                    or current_article.affiliations
                    or current_article.presentation_abstract
                ):
                    if not current_article.session_name:
                        if previous_article is not None:
                            previous_article.session_title += (
                                current_article.session_title
                            )
                            previous_article.affiliations.extend(
                                current_article.affiliations
                            )
                            previous_article.location.extend(current_article.location)
                            previous_article.presentation_abstract += (
                                current_article.presentation_abstract
                            )
                            current_article = previous_article
                    articles.append(current_article)
                    previous_article = current_article
                    current_article = Article()

    file.close()

    return articles


def update_excel_file(article):
    workbook = openpyxl.load_workbook(xl_file)
    sheet = workbook.active

    row = sheet.max_row + 1

    existing_authors = (
        set(sheet.cell(row=row, column=1).value.split(", "))
        if sheet.cell(row=row, column=1).value
        else set()
    )

    for author in article.authors:
        if not author:
            continue

        if len(author) == 1:
            if existing_authors:
                previous_author = existing_authors.pop()
                author = previous_author + " " + author
            else:
                continue

        if author.count(",") > 2:
            continue

        if author.startswith(", "):
            author = author[2:]

        sheet.cell(row=row, column=1).value = author
        sheet.cell(row=row, column=2).value = ", ".join(article.affiliations)
        sheet.cell(row=row, column=3).value = ", ".join(article.location)
        sheet.cell(row=row, column=4).value = article.session_name
        sheet.cell(row=row, column=5).value = article.session_title
        sheet.cell(row=row, column=6).value = article.presentation_abstract

        font = Font(name="Arial", size=10)
        for col in range(1, 7):
            sheet.cell(row=row, column=col).font = font

        existing_authors.add(author)

    workbook.save(xl_file)


def read_existing_data():
    workbook = openpyxl.load_workbook(xl_file)
    sheet = workbook.active

    existing_data = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        authors = row[0]
        affiliations = row[1]
        location = row[2]
        session_name = row[3]
        session_title = row[4]
        presentation_abstract = row[5]

        existing_data.add(
            (
                authors,
                affiliations,
                location,
                session_name,
                session_title,
                presentation_abstract,
            )
        )

    return existing_data


def process_articles(articles):
    existing_data = read_existing_data()

    for article in articles:
        for author in article.authors:
            article_copy = Article()
            article_copy.session_name = article.session_name
            article_copy.session_title = article.session_title
            article_copy.authors = [author]
            article_copy.affiliations = article.affiliations.copy()
            article_copy.location = article.location.copy()
            article_copy.presentation_abstract = article.presentation_abstract

            if article_copy.authors:
                if (
                    ", ".join(article_copy.authors),
                    ", ".join(article_copy.affiliations),
                    ", ".join(article_copy.location),
                    article_copy.session_name,
                    article_copy.session_title,
                    article_copy.presentation_abstract,
                ) not in existing_data:
                    update_excel_file(article_copy)
                    existing_data.add(
                        (
                            ", ".join(article_copy.authors),
                            ", ".join(article_copy.affiliations),
                            ", ".join(article_copy.location),
                            article_copy.session_name,
                            article_copy.session_title,
                            article_copy.presentation_abstract,
                        )
                    )


if __name__ == "__main__":
    pdf_articles = extract_articles_from_pdf(pdf_file)

    pdf_articles = [
        article
        for article in pdf_articles
        if any(author.strip() for author in article.authors)
    ]

    process_articles(pdf_articles)
